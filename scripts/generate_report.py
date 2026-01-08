#!/usr/bin/env python3
"""
Generate comprehensive Excel report from database.
Exports all tracking data to an Excel file for analysis.
"""
import os
import sys
import json
from datetime import datetime
from pathlib import Path

# Add project root to path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not installed. Install with: pip install openpyxl")


def generate_excel_report(db_path: str = None, output_path: str = None):
    """
    Generate comprehensive Excel report from database.
    
    Args:
        db_path: Path to SQLite database
        output_path: Path for output Excel file
    """
    if not OPENPYXL_AVAILABLE:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        return None
    
    # Default paths
    if db_path is None:
        db_path = "./feedback_data/intellipest.db"
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"./reports/intellipest_report_{timestamp}.xlsx"
    
    # Ensure output directory exists
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    
    # Connect to database
    import sqlite3
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def style_header_row(ws, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
    
    def auto_column_width(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # ==================== Sheet 1: Summary Dashboard ====================
    ws_summary = wb.active
    ws_summary.title = "Summary Dashboard"
    
    # Get summary stats
    cursor.execute("SELECT COUNT(*) FROM users")
    total_users = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM feedback_entries")
    total_feedback = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM training_runs WHERE status = 'completed'")
    completed_runs = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM archived_images")
    archived_images = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(*) FROM scheduler_checks WHERE training_triggered = 1")
    triggered_count = cursor.fetchone()[0]
    
    summary_data = [
        ["INTELLI-PEST SYSTEM REPORT", ""],
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["", ""],
        ["SYSTEM STATISTICS", ""],
        ["Total Users", total_users],
        ["Total Feedback Entries", total_feedback],
        ["Completed Training Runs", completed_runs],
        ["Archived Images", archived_images],
        ["Auto-Triggered Trainings", triggered_count],
        ["", ""],
        ["TRAINING THRESHOLDS", ""],
        ["Fine-Tuning: Images per Class", 10],
        ["Fine-Tuning: Total Images", 150],
        ["Fine-Tuning: Min Classes", 3],
        ["Comprehensive: Total Historical", 1000],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx in [1, 4, 11]:
                cell.font = Font(bold=True, size=14)
            cell.border = border
    
    auto_column_width(ws_summary)
    
    # ==================== Sheet 2: Users ====================
    ws_users = wb.create_sheet("Users")
    
    cursor.execute("""
        SELECT user_id, email, user_type, total_submissions, total_feedbacks,
               correct_feedbacks, correction_feedbacks, trust_score, is_flagged,
               first_seen, last_seen
        FROM users ORDER BY last_seen DESC
    """)
    users = cursor.fetchall()
    
    user_headers = ["User ID", "Email", "Type", "Submissions", "Feedbacks", 
                    "Correct", "Corrections", "Trust Score", "Flagged", 
                    "First Seen", "Last Seen"]
    for col, header in enumerate(user_headers, 1):
        ws_users.cell(row=1, column=col, value=header)
    style_header_row(ws_users, len(user_headers))
    
    for row_idx, user in enumerate(users, 2):
        ws_users.cell(row=row_idx, column=1, value=user["user_id"])
        ws_users.cell(row=row_idx, column=2, value=user["email"])
        ws_users.cell(row=row_idx, column=3, value=user["user_type"])
        ws_users.cell(row=row_idx, column=4, value=user["total_submissions"])
        ws_users.cell(row=row_idx, column=5, value=user["total_feedbacks"])
        ws_users.cell(row=row_idx, column=6, value=user["correct_feedbacks"])
        ws_users.cell(row=row_idx, column=7, value=user["correction_feedbacks"])
        ws_users.cell(row=row_idx, column=8, value=round(user["trust_score"], 2))
        ws_users.cell(row=row_idx, column=9, value="Yes" if user["is_flagged"] else "No")
        ws_users.cell(row=row_idx, column=10, value=user["first_seen"])
        ws_users.cell(row=row_idx, column=11, value=user["last_seen"])
        
        for col in range(1, len(user_headers) + 1):
            ws_users.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_users)
    
    # ==================== Sheet 3: Training Runs ====================
    ws_training = wb.create_sheet("Training Runs")
    
    cursor.execute("""
        SELECT run_id, training_type, status, started_at, completed_at,
               model_version, epochs_planned, epochs_completed, total_images_used,
               best_upright_accuracy, best_rotation_accuracy, 
               early_stopped, collapse_detected, training_duration_seconds, error_message
        FROM training_runs ORDER BY created_at DESC
    """)
    runs = cursor.fetchall()
    
    training_headers = ["Run ID", "Type", "Status", "Started", "Completed",
                        "Model Version", "Epochs Plan", "Epochs Done", "Images Used",
                        "Best Upright %", "Best Rotation %", "Early Stopped", 
                        "Collapse", "Duration (s)", "Error"]
    for col, header in enumerate(training_headers, 1):
        ws_training.cell(row=1, column=col, value=header)
    style_header_row(ws_training, len(training_headers))
    
    for row_idx, run in enumerate(runs, 2):
        ws_training.cell(row=row_idx, column=1, value=run["run_id"])
        ws_training.cell(row=row_idx, column=2, value=run["training_type"])
        cell = ws_training.cell(row=row_idx, column=3, value=run["status"])
        if run["status"] == "completed":
            cell.fill = success_fill
        elif run["status"] == "failed":
            cell.fill = error_fill
        elif run["status"] == "running":
            cell.fill = warning_fill
        ws_training.cell(row=row_idx, column=4, value=run["started_at"])
        ws_training.cell(row=row_idx, column=5, value=run["completed_at"])
        ws_training.cell(row=row_idx, column=6, value=run["model_version"])
        ws_training.cell(row=row_idx, column=7, value=run["epochs_planned"])
        ws_training.cell(row=row_idx, column=8, value=run["epochs_completed"])
        ws_training.cell(row=row_idx, column=9, value=run["total_images_used"])
        ws_training.cell(row=row_idx, column=10, value=round(run["best_upright_accuracy"] or 0, 2))
        ws_training.cell(row=row_idx, column=11, value=round(run["best_rotation_accuracy"] or 0, 2))
        ws_training.cell(row=row_idx, column=12, value="Yes" if run["early_stopped"] else "No")
        ws_training.cell(row=row_idx, column=13, value="Yes" if run["collapse_detected"] else "No")
        ws_training.cell(row=row_idx, column=14, value=round(run["training_duration_seconds"] or 0, 1))
        ws_training.cell(row=row_idx, column=15, value=run["error_message"])
        
        for col in range(1, len(training_headers) + 1):
            ws_training.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_training)
    
    # ==================== Sheet 4: Scheduler Checks ====================
    ws_scheduler = wb.create_sheet("Scheduler Checks")
    
    cursor.execute("""
        SELECT check_timestamp, total_feedback_images, classes_with_images,
               threshold_met, comprehensive_threshold_met, training_triggered,
               training_type, reason
        FROM scheduler_checks ORDER BY check_timestamp DESC LIMIT 500
    """)
    checks = cursor.fetchall()
    
    scheduler_headers = ["Timestamp", "Total Images", "Classes with Images", 
                         "Threshold Met", "Comprehensive Met", "Triggered",
                         "Training Type", "Reason"]
    for col, header in enumerate(scheduler_headers, 1):
        ws_scheduler.cell(row=1, column=col, value=header)
    style_header_row(ws_scheduler, len(scheduler_headers))
    
    for row_idx, check in enumerate(checks, 2):
        ws_scheduler.cell(row=row_idx, column=1, value=check["check_timestamp"])
        ws_scheduler.cell(row=row_idx, column=2, value=check["total_feedback_images"])
        ws_scheduler.cell(row=row_idx, column=3, value=check["classes_with_images"])
        ws_scheduler.cell(row=row_idx, column=4, value="Yes" if check["threshold_met"] else "No")
        ws_scheduler.cell(row=row_idx, column=5, value="Yes" if check["comprehensive_threshold_met"] else "No")
        cell = ws_scheduler.cell(row=row_idx, column=6, value="Yes" if check["training_triggered"] else "No")
        if check["training_triggered"]:
            cell.fill = success_fill
        ws_scheduler.cell(row=row_idx, column=7, value=check["training_type"])
        ws_scheduler.cell(row=row_idx, column=8, value=check["reason"])
        
        for col in range(1, len(scheduler_headers) + 1):
            ws_scheduler.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_scheduler)
    
    # ==================== Sheet 5: System Events ====================
    ws_events = wb.create_sheet("System Events")
    
    cursor.execute("""
        SELECT created_at, event_type, component, severity, message
        FROM system_events ORDER BY created_at DESC LIMIT 1000
    """)
    events = cursor.fetchall()
    
    event_headers = ["Timestamp", "Event Type", "Component", "Severity", "Message"]
    for col, header in enumerate(event_headers, 1):
        ws_events.cell(row=1, column=col, value=header)
    style_header_row(ws_events, len(event_headers))
    
    for row_idx, event in enumerate(events, 2):
        ws_events.cell(row=row_idx, column=1, value=event["created_at"])
        ws_events.cell(row=row_idx, column=2, value=event["event_type"])
        ws_events.cell(row=row_idx, column=3, value=event["component"])
        cell = ws_events.cell(row=row_idx, column=4, value=event["severity"])
        if event["severity"] == "error":
            cell.fill = error_fill
        elif event["severity"] == "warning":
            cell.fill = warning_fill
        ws_events.cell(row=row_idx, column=5, value=event["message"])
        
        for col in range(1, len(event_headers) + 1):
            ws_events.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_events)
    
    # ==================== Sheet 6: Archived Images ====================
    ws_archived = wb.create_sheet("Archived Images")
    
    cursor.execute("""
        SELECT image_hash, original_path, archive_path, training_run_id,
               training_type, class_name, source_folder, archived_at
        FROM archived_images ORDER BY archived_at DESC LIMIT 1000
    """)
    archived = cursor.fetchall()
    
    archived_headers = ["Image Hash", "Original Path", "Archive Path", "Training Run",
                        "Training Type", "Class", "Source Folder", "Archived At"]
    for col, header in enumerate(archived_headers, 1):
        ws_archived.cell(row=1, column=col, value=header)
    style_header_row(ws_archived, len(archived_headers))
    
    for row_idx, img in enumerate(archived, 2):
        ws_archived.cell(row=row_idx, column=1, value=img["image_hash"])
        ws_archived.cell(row=row_idx, column=2, value=img["original_path"])
        ws_archived.cell(row=row_idx, column=3, value=img["archive_path"])
        ws_archived.cell(row=row_idx, column=4, value=img["training_run_id"])
        ws_archived.cell(row=row_idx, column=5, value=img["training_type"])
        ws_archived.cell(row=row_idx, column=6, value=img["class_name"])
        ws_archived.cell(row=row_idx, column=7, value=img["source_folder"])
        ws_archived.cell(row=row_idx, column=8, value=img["archived_at"])
        
        for col in range(1, len(archived_headers) + 1):
            ws_archived.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_archived)
    
    # ==================== Sheet 7: Feedback Entries ====================
    ws_feedback = wb.create_sheet("Feedback Entries")
    
    cursor.execute("""
        SELECT feedback_id, user_id, image_hash, predicted_class, confidence,
               is_correct, corrected_class, submitted_at
        FROM feedback_entries ORDER BY submitted_at DESC LIMIT 1000
    """)
    feedback = cursor.fetchall()
    
    feedback_headers = ["Feedback ID", "User ID", "Image Hash", "Predicted Class",
                        "Confidence", "Is Correct", "Corrected Class", "Submitted At"]
    for col, header in enumerate(feedback_headers, 1):
        ws_feedback.cell(row=1, column=col, value=header)
    style_header_row(ws_feedback, len(feedback_headers))
    
    for row_idx, fb in enumerate(feedback, 2):
        ws_feedback.cell(row=row_idx, column=1, value=fb["feedback_id"])
        ws_feedback.cell(row=row_idx, column=2, value=fb["user_id"])
        ws_feedback.cell(row=row_idx, column=3, value=fb["image_hash"])
        ws_feedback.cell(row=row_idx, column=4, value=fb["predicted_class"])
        ws_feedback.cell(row=row_idx, column=5, value=round(fb["confidence"] or 0, 4))
        cell = ws_feedback.cell(row=row_idx, column=6, value="Yes" if fb["is_correct"] else "No")
        if fb["is_correct"]:
            cell.fill = success_fill
        else:
            cell.fill = warning_fill
        ws_feedback.cell(row=row_idx, column=7, value=fb["corrected_class"])
        ws_feedback.cell(row=row_idx, column=8, value=fb["submitted_at"])
        
        for col in range(1, len(feedback_headers) + 1):
            ws_feedback.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_feedback)
    
    # ==================== Sheet 8: Audit Log ====================
    ws_audit = wb.create_sheet("Audit Log")
    
    cursor.execute("""
        SELECT created_at, action, entity_type, entity_id, old_value, new_value, notes
        FROM audit_log ORDER BY created_at DESC LIMIT 500
    """)
    audit = cursor.fetchall()
    
    audit_headers = ["Timestamp", "Action", "Entity Type", "Entity ID", 
                     "Old Value", "New Value", "Notes"]
    for col, header in enumerate(audit_headers, 1):
        ws_audit.cell(row=1, column=col, value=header)
    style_header_row(ws_audit, len(audit_headers))
    
    for row_idx, entry in enumerate(audit, 2):
        ws_audit.cell(row=row_idx, column=1, value=entry["created_at"])
        ws_audit.cell(row=row_idx, column=2, value=entry["action"])
        ws_audit.cell(row=row_idx, column=3, value=entry["entity_type"])
        ws_audit.cell(row=row_idx, column=4, value=entry["entity_id"])
        ws_audit.cell(row=row_idx, column=5, value=entry["old_value"])
        ws_audit.cell(row=row_idx, column=6, value=entry["new_value"])
        ws_audit.cell(row=row_idx, column=7, value=entry["notes"])
        
        for col in range(1, len(audit_headers) + 1):
            ws_audit.cell(row=row_idx, column=col).border = border
    
    auto_column_width(ws_audit)
    
    # Close database
    conn.close()
    
    # Save workbook
    wb.save(output_path)
    print(f"✅ Report generated: {output_path}")
    
    return output_path


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Excel report from database")
    parser.add_argument("--db", default="./feedback_data/intellipest.db", 
                        help="Path to SQLite database")
    parser.add_argument("--output", default=None,
                        help="Output Excel file path")
    
    args = parser.parse_args()
    
    generate_excel_report(db_path=args.db, output_path=args.output)
